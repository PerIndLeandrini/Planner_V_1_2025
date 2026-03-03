import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

# -----------------------------
# AUTH LOGIN
# -----------------------------
# -----------------------------
# AUTH LOGIN (UI moderna)
# -----------------------------
def check_login():

    if "auth_ok" not in st.session_state:
        st.session_state.auth_ok = False

    if st.session_state.auth_ok:
        return True

    # ===== CSS CUSTOM =====
    st.markdown("""
        <style>
        .login-box {
            max-width: 420px;
            margin: auto;
            padding: 35px;
            border-radius: 14px;
            background-color: #ffffff;
            box-shadow: 0 6px 18px rgba(0,0,0,0.08);
            text-align: center;
        }

        .login-title {
            font-size: 26px;
            font-weight: 600;
            margin-bottom: 10px;
        }

        .login-sub {
            color: #666;
            margin-bottom: 25px;
            font-size: 14px;
        }

        div[data-testid="stTextInput"] input {
            text-align: center;
        }

        .logo-img {
            width: 70px;
            margin-bottom: 10px;
        }
        </style>
    """, unsafe_allow_html=True)

    # spazio verticale
    st.write("")
    st.write("")

    # ===== CARD LOGIN =====
    col1, col2, col3 = st.columns([1,2,1])

    with col2:
        st.markdown('<div class="login-box">', unsafe_allow_html=True)

        # 🔵 LOGO (metti il tuo file nella repo)
        st.image("logo.png", width=70)   # <-- cambia nome se vuoi

        st.markdown('<div class="login-title">Accesso Orderbook</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-sub">Inserisci le credenziali per continuare</div>', unsafe_allow_html=True)

        username = st.text_input("Utente", label_visibility="collapsed", placeholder="👤 Utente")
        password = st.text_input("Password", type="password",
                                 label_visibility="collapsed",
                                 placeholder="🔒 Password")

        if st.button("Accedi", use_container_width=True):

            users = st.secrets.get("auth", {}).get("users", [])
            pwds  = st.secrets.get("auth", {}).get("passwords", [])

            if username in users:
                idx = users.index(username)

                if idx < len(pwds) and password == pwds[idx]:
                    st.session_state.auth_ok = True
                    st.rerun()
                else:
                    st.error("Password errata")
            else:
                st.error("Utente non valido")

        st.markdown("</div>", unsafe_allow_html=True)

    st.stop()
    
st.set_page_config(page_title="Orderbook Generator", layout="wide")
check_login()
st.title("📘 Orderbook – Generatore da CSV")

st.caption(
    "Carica il CSV (compilato) + il template Orderbook vuoto (solo intestazioni). "
    "Ottieni l'Excel compilato e una preview online."
)

# -----------------------------
# Helpers
# -----------------------------
def read_csv_flexible(uploaded_file) -> pd.DataFrame:
    """
    Legge CSV in modo tollerante:
    - prova con ; poi con ,
    - forza dtype=str
    """
    raw = uploaded_file.getvalue()

    # prova separatore ;
    try:
        df = pd.read_csv(BytesIO(raw), sep=";", dtype=str).fillna("")
        if df.shape[1] > 1:
            return df
    except Exception:
        pass

    # fallback separatore ,
    df = pd.read_csv(BytesIO(raw), sep=",", dtype=str).fillna("")
    return df

import re
from datetime import date

DATE_FMT_ITA = "%d/%m/%Y"

def try_parse_date(val):
    """
    Ritorna un datetime.date se val è una data riconoscibile, altrimenti None.
    Gestisce:
    - 'dd/mm/yyyy'
    - 'yyyy-mm-dd'
    - 'yyyy-mm-dd hh:mm:ss'
    """
    if val is None:
        return None

    s = str(val).strip()
    if not s:
        return None

    # Caso già in formato ITA dd/mm/yyyy
    try:
        dt = datetime.strptime(s[:10], DATE_FMT_ITA)
        return dt.date()
    except Exception:
        pass

    # Caso ISO yyyy-mm-dd (con o senza ora)
    # es: 2026-06-27 oppure 2026-06-27 00:00:00
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.isna(dt):
                return None
            return dt.date()
        except Exception:
            return None

    return None

def fill_template_from_df(template_bytes: bytes, df: pd.DataFrame, sheet_index: int = 0) -> bytes:
    """
    Scrive il contenuto del CSV nel template Excel:
    - lascia intatta la riga 1 (intestazioni template)
    - scrive i dati da A2 in poi
    - mantiene formato date italiano
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.worksheets[sheet_index]

    # pulisce eventuali righe sotto header
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    start_row = 2
    start_col = 1  # colonna A

    for i, row in df.iterrows():
        for j, val in enumerate(row, start=start_col):

            cell = ws.cell(row=start_row + i, column=j)

            # --- gestione date ---
            d = try_parse_date(val)

            if d is not None:
                cell.value = d
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = "" if pd.isna(val) else str(val)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------
# UI
# -----------------------------
c1, c2 = st.columns(2)
with c1:
    up_csv = st.file_uploader("📤 Carica CSV orderbook (compilato)", type=["csv"], key="csv")
with c2:
    up_tpl = st.file_uploader("📤 Carica template Orderbook vuoto (.xlsx)", type=["xlsx"], key="tpl")

if not up_csv or not up_tpl:
    st.info("Carica entrambi i file per procedere.")
    st.stop()

# Leggi CSV
try:
    df = read_csv_flexible(up_csv)
except Exception as e:
    st.error(f"Errore lettura CSV: {e}")
    st.stop()

st.divider()

if st.button("🚀 Genera Orderbook Excel compilato", use_container_width=True):
    try:
        out_bytes = fill_template_from_df(up_tpl.getvalue(), df, sheet_index=0)
    except Exception as e:
        st.error(f"Errore generazione Excel: {e}")
        st.stop()

    fname = f"orderbook_compilato_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    st.success("✅ Excel generato.")
    st.download_button(
        "⬇️ Scarica Orderbook compilato",
        data=out_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )





